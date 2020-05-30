import openpyxl as oxl
from string import ascii_letters,digits
from random import randint,choice
class Xlrw:
    def __init__(self,path):
      self.path = path
      
    def read(self):
        l=[]
        xl=oxl.load_workbook(self.path)
        sh=xl.active
        [l.append((sh.cell(row=i,column=1)).value) for i in range(1,sh.max_row+1)]
        xl.close()
        return l

    def passGen(self,maillist):
        mpdict={}
        [mpdict.update({mail:("".join(choice(ascii_letters+digits)for x in range(randint(6,6))))}) for mail in maillist]
        return mpdict
    
    def passWrite(self,mpdict):
        xl=oxl.load_workbook(self.path)
        sh=xl.active
        for x,i in zip(mpdict.values(),range(1,len(mpdict)+1)):
            (sh.cell(row=i,column=2)).value=x
        xl.save(self.path)
        xl.close()




