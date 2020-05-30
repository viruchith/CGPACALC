import psycopg2
from flask_bcrypt import Bcrypt
import openpyxl as oxl
from AppCredentials import Credentials as creds

bcrypt=Bcrypt()

connection = psycopg2.connect(user=creds.DB_USERNAME,
                              password=creds.DB_PASSWORD,
                              host="localhost",
                              port="5432",
                              database=creds.DB_NAME)
cursor=connection.cursor()
xl=oxl.load_workbook("/Users/ganesankoundappan/Projects/CSEwebApp/<dasdasd>.xlsx")
sh=xl.active
for i in range(1,sh.max_row+1):
    admission_no=(sh.cell(row=i,column=6).value)
    email_id=(sh.cell(row=i,column=1).value)
    password=(sh.cell(row=i,column=2).value)
    name=(sh.cell(row=i,column=3).value)
    mobile = (sh.cell(row=i, column=5).value)

    hashed = bcrypt.generate_password_hash(password).decode('utf-8')
    cursor.execute("""INSERT INTO student_primary_credentials (admission_no,email_id,password,student_name,mobile_no) VALUES (%s,%s,%s,%s,%s); """,(admission_no, email_id, hashed, name, mobile))
    
connection.commit()
connection.close()    

